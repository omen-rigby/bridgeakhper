import shutil
import transliterate
import datetime
from jinja2 import Template
from urllib.parse import urlparse
from openpyxl.styles import PatternFill
from result_getter import ResultGetter
from generate import generate
from movement import Movement
from players import *
from inline_key import *
from functools import wraps
from config import init_config
from swiss import SwissMovement
from match_handlers import MatchHandlers
from util import decorate_all_functions, revert_name
from openpyxl import Workbook
from exceptions import MovementError

CHANGE_FLOWS = ('update_player', 'add_player', 'remove_player', 'view_board', 'remove_board', 'remove pair',
                'tourney_coeff', 'tournament_title', 'rounds', 'add_td', 'config_update', 'penalty', 'table_card', 'move_card',
                'select_session', 'load_db', 'howell', 'mitchell', 'barometer', 'xlsx')
ALLOWED_IN_GROUP = ('/end', '/movecards', '/startround', '/restartswiss', '/endround')


def command_eligibility(func):
    @wraps(func)
    def wrapper(update: Update, context: CallbackContext):
        if not update.message or not update.message.text:
            # reaction or edited message
            return
        if CONFIG["city"] and update.message and update.message.text and update.message.text.startswith('/'):
            for key in CHANGE_FLOWS:
                context.user_data[key] = False
            context.user_data['names'] = None
        if update.effective_chat.id < 0 and update.message.text.startswith('/') and \
                update.message.text not in ALLOWED_IN_GROUP:
            send(chat_id=update.effective_chat.id, text="This bot shouldn't be called in groups", context=context)
            context.bot.deleteMessage(chat_id=update.effective_chat.id, message_id=update.message.message_id)
            raise Exception("Bad command")
        return func(update, context)
    return wrapper


@decorate_all_functions(command_eligibility)
class CommandHandlers:
    @staticmethod
    def start(update: Update, context: CallbackContext):
        if CONFIG["city"]:
            send(chat_id=update.effective_chat.id,
                 text="Started bridge scorer. What do you want to do?",
                 reply_buttons=("/board",),
                 context=context)
        else:
            send(chat_id=update.effective_chat.id,
                 text="Started BWS aggregator. Drag files to upload",
                 context=context)

    @staticmethod
    def clear_all(update: Update, context: CallbackContext):
        if is_director(update):
            CommandHandlers.clear_db(update, context)
        else:
            send(chat_id=update.effective_chat.id,
                 text="Not enough permissions",
                 context=context)

    @staticmethod
    def start_multi_session(update: Update, context: CallbackContext):
        """
        Players & boards for session 2 are numbered 101..., etc.
        """
        context.bot_data["current_session"] = -1
        context.user_data["current_session"] = -1
        send(chat_id=update.effective_chat.id,
             text="Started multisession mode. "
                  "Use /session to start first session or /endmultisession to return to single session.",
             reply_buttons=("/session", "/endmultisession"),
             context=context)

    @staticmethod
    def switch_session(update: Update, context: CallbackContext):
        if context.user_data.get("select_session"):
            context.user_data['current_session'] = int(update.message.text) - 1
            context.user_data["select_session"] = False
        else:
            reply_buttons = [str(b + 1) for b in range(context.bot_data.get('current_session'))
                             if b != context.user_data.get('current_session')]
            send(chat_id=update.effective_chat.id,
                 text="Select session to edit",
                 reply_buttons=reply_buttons,
                 context=context)
            context.user_data["select_session"] = True

    @staticmethod
    def end_multi_session(update: Update, context: CallbackContext):
        context.bot_data["current_session"] = None
        context.user_data["current_session"] = None
        del context.bot_data["current_session"]
        del context.user_data["current_session"]
        send(chat_id=update.effective_chat.id,
             text="Exited multi-session mode",
             context=context)

    @staticmethod
    def config(update: Update, context: CallbackContext):
        conn = TourneyDB.connect()
        cursor = conn.cursor()
        cursor.execute("select key,value from config order by key")

        if is_director(update):
            string_config = '\n'.join(f'{k} = {v}' for (k, v) in cursor.fetchall())
            send(chat_id=update.effective_chat.id,
                 text=f"Current global parameters:\n{string_config}",
                 context=context)
        else:
            send(chat_id=update.effective_chat.id,
                 text="Not enough permissions",
                 context=context)
        conn.close()

    @staticmethod
    def config_update(update: Update, context: CallbackContext):
        if is_director(update):
            if context.user_data.get("config_update"):
                try:
                    k, v = update.message.text.split('=')
                    k = k.strip()
                    v = v.strip()
                except (IndexError, ValueError):
                    return send(chat_id=update.effective_chat.id,
                                text="Incorrect invocation, try mp:neuberg=false",
                                context=context)
                conn = TourneyDB.connect()
                cursor = conn.cursor()
                cursor.execute("select key from config order by key")
                keys = [c[0] for c in cursor.fetchall()]
                if k not in keys:
                    cursor.execute(f"insert into config (key, value) values ('{k}', '{v}')")
                else:
                    cursor.execute(f"update config set value='{v}' where key='{k}'")
                send(chat_id=update.effective_chat.id,
                     text=f"{k} is set to {v}",
                     context=context)
                cursor.execute("select key,value from config order by key")
                string_config = '\n'.join(f'{k} = {v}' for (k, v) in cursor.fetchall())
                send(chat_id=update.effective_chat.id,
                     text=f"Current global parameters:\n{string_config}",
                     context=context)
                conn.commit()
                init_config()
                context.user_data["config_update"] = False
                conn.close()
            else:
                send(chat_id=update.effective_chat.id,
                     text=f"""Type in the desired parameter to update, e.g. imp:fractional=true.""",
                     context=context)
                context.user_data["config_update"] = True
        else:
            send(chat_id=update.effective_chat.id,
                 text="Not enough permissions",
                 context=context)

    @staticmethod
    def board(update: Update, context: CallbackContext):
        board_numbers(update, context)

    @staticmethod
    def init(update: Update, context: CallbackContext):
        try:
            conn = TourneyDB.connect()
        except Exception:
            send(chat_id=update.effective_chat.id,
                 text="Couldn't connect to tournament database.",
                 context=context)
            raise
        try:
            cursor = conn.cursor()
        except:
            send(chat_id=update.effective_chat.id,
                 text="Couldn't connect to tournament database.",
                 context=context)
            conn.close()
            raise

        if 'Swiss' not in CONFIG['scoring'] and context.bot_data.get('current_session') is not None:
            context.bot_data['current_session'] += 1
            context.user_data['current_session'] = context.bot_data['current_session']
        if 'rounds' in CONFIG.keys():
            del CONFIG['rounds']
        context.bot_data["maxboard"] = 0
        context.bot_data["maxpair"] = 0
        context.user_data["names"] = None
        context.bot_data["movement"] = None
        context.bot_data["allowed_boards"] = {}
        for key in CHANGE_FLOWS:
            context.user_data[key] = False
        initial_config = json.load(open(os.path.abspath(__file__).replace(os.path.basename(__file__), "config.json")))
        if context.bot_data.get('current_session') is None or context.bot_data.get('current_session') == 0:
            CONFIG["tournament_title"] = initial_config["tournament_title"]
            CONFIG["tourney_coeff"] = 0.25
        init_config()
        if (context.bot_data.get('current_session') or 0) > 0:
            session_index = context.bot_data.get('current_session')
            cursor.execute('select number,partnership,rank,rank_ru from names')
            all_names = cursor.fetchall()
            current_session_names = [n for n in all_names if 100 * session_index < n[0] < 100 * session_index + 99]
            previous_session_names = [n for n in all_names if -100 + 100 * session_index < n[0] < 100 * session_index]
            if not current_session_names:
                for n in previous_session_names:
                    cursor.execute(f"""insert into names ("number", partnership, "rank", rank_ru) 
                        VALUES({n[0] + 100 * session_index}, '{n[1]}', {n[2]}, {n[3]})""")
            cursor.execute(f'select distinct "number" from protocols where 100 * {session_index - 1} < "number" and "number" < 100 * {session_index} order by number')
            boards_previous = [b[0] for b in cursor.fetchall()]
            last_board = [b for b in boards_previous if b + 1 not in boards_previous][0] % 100
            if modulo := CONFIG.get('sessions', {}).get('modulo'):
                context.bot_data["first_board"] = last_board % modulo + 1
            send(chat_id=update.effective_chat.id,
                 text=f"Started session {context.bot_data['current_session'] + 1}.\nEnter the number of boards",
                 reply_buttons=[],
                 context=context)
            cursor.execute(f"update session set session_index={context.bot_data['current_session']}")
        else:
            send(chat_id=update.effective_chat.id,
                 text="Started session. Enter scoring",
                 reply_buttons=["MPs", "IMPs", "Cross-IMPs", "Swiss IMPs"],
                 context=context)
            initiator = update.message.from_user.username or update.message.from_user.id
            session_index = context.bot_data["current_session"]
            cursor.execute('select COUNT(*) from session')
            session_found = cursor.fetchone()
            if session_found[0]:
                cursor.execute(f"update session set initiator='{initiator}',boards=NULL,pairs=NULL,scoring=NULL,session_index={session_index}")
            else:
                cursor.execute(
                    f"""insert into session (pairs, boards, scoring, initiator,session_index) VALUES (NULL,NULL,NULL,'{initiator}',{session_index})""")
        conn.commit()

        if update.effective_chat.id != BITKIN_ID:
            send(chat_id=BITKIN_ID,
                 text=f"Session started by {initiator}",
                 reply_buttons=[],
                 context=context)

    @staticmethod
    def reuse_names(update: Update, context: CallbackContext):
        TourneyDB.clear_tables(("boards", "protocols"))
        send(chat_id=update.effective_chat.id,
             text=f"""Restarting session saving names""",
             reply_buttons=[],
             context=context)
        CommandHandlers.init(update, context)

    @staticmethod
    def clear_db(update: Update, context: CallbackContext):
        TourneyDB.clear_tables()
        send(chat_id=update.effective_chat.id,
             text=f"""Starting new session""",
             reply_buttons=[],
             context=context)
        CommandHandlers.init(update, context)

    @staticmethod
    def start_session(update: Update, context: CallbackContext):
        if is_director(update):
            if db_path.startswith('postgres:'):
                conn = TourneyDB.connect()
                cursor = conn.cursor()
                first = 100 * current_session(context)
                if context.bot_data.get('current_session', -1) == -1:
                    for table in ('boards', 'protocols', 'names'):
                        cursor.execute(f'select * from {table} where {first} < number and number < {first + 100} LIMIT 1')
                        if cursor.fetchone():
                            break
                    else:
                        conn.close()
                        return CommandHandlers.init(update, context)
                    conn.close()
                    return send(chat_id=update.effective_chat.id,
                                text="Session exists in database. Clear and start new tournament or reuse the existing data?",
                                reply_buttons=["Clear", "Reuse", "Reuse names"],
                                context=context)
            else:
                generate()
            CommandHandlers.init(update, context)
        else:
            return CommandHandlers.missing(update, context)

    @staticmethod
    def missing(update, context):
        conn = TourneyDB.connect()
        boards_num = context.bot_data.get("maxboard")
        pairs_num = context.bot_data.get("maxpair")
        cursor = conn.cursor()
        first = 100 * current_session(context)
        if 'wiss' in CONFIG['scoring']:
            boards_per_round = context.bot_data['maxboard'] // CONFIG['rounds']
            round_number = context.bot_data['movement'].round
            max_board = boards_per_round * round_number + 1
        else:
            max_board = boards_num + 1
        cursor.execute(f'select * from boards where {first} < number and number < {first + max_board}')
        submitted_boards = [b[0] for b in cursor.fetchall()]
        boards = ", ".join([str(b) for b in range(first + 1, first + max_board) if b not in submitted_boards])
        cursor.execute(f'select * from protocols where {first} < number and number < {first + max_board}')
        protocols = list(set(cursor.fetchall()))
        expected_records = boards_num * (pairs_num // 2)
        boards_with_missing_results = ", ".join([str(i) for i in range(first + 1, first + max_board)
                                                 if len([p for p in protocols if p[0] == i]) < pairs_num // 2])
        cursor.execute(f'select * from names where {first} < number and number < {first + 100}')
        names = len(cursor.fetchall())
        if boards_num and pairs_num:
            send(chat_id=update.effective_chat.id,
                 text=f"""Active session: {pairs_num} pairs {CONFIG['scoring']}
Missing boards: {boards}
Missing results for boards: {boards_with_missing_results}
Submitted {len(protocols)}/{expected_records} results
Submitted {names} names""",
                 reply_buttons=[],
                 context=context)
        else:
            send(chat_id=update.effective_chat.id,
                 text=f"""Session not started yet""",
                 reply_buttons=[],
                 context=context)
        conn.close()

    @staticmethod
    def scoring(update: Update, context: CallbackContext):
        CONFIG["scoring"] = update.message.text
        if is_director(update):
            send(chat_id=update.effective_chat.id,
                 text="Enter number of boards",
                 reply_buttons=[],
                 context=context)

    @staticmethod
    def names(update: Update, context: CallbackContext):
        context.user_data["names"] = 0
        conn = TourneyDB.connect()
        cursor = conn.cursor()
        first = 100 * current_session(context)
        cursor.execute(f"Select number from names where {first} < number and number < {first + 100}")
        added = list(set([c[0] - first for c in cursor.fetchall()]))
        conn.close()
        all_pairs = get_valid_pairs(context)
        buttons = [l for l in all_pairs if l not in added]
        send(chat_id=update.effective_chat.id,
             text="Enter pair number",
             reply_buttons=buttons,
             context=context)

    @staticmethod
    def players(update: Update, context: CallbackContext):
        conn = TourneyDB.connect()
        cursor = conn.cursor()
        first = 100 * current_session(context)
        cursor.execute(f"Select number,partnership from names where {first} < number and number < {first + 100} "
                       f"order by number")
        found_pairs = [f'{p[0]}: {p[1]}' for p in cursor.fetchall()]
        send(chat_id=update.effective_chat.id,
             text='\n'.join(found_pairs),
             context=context)
        conn.close()

    @staticmethod
    def move_card(update: Update, context: CallbackContext):
        if not context.bot_data.get("movement"):
            send(chat_id=update.effective_chat.id,
                 text="Movement not set",
                 context=context)
        elif context.user_data.get("move_card"):
            context.user_data["move_card"] = False
            send(chat_id=update.effective_chat.id,
                 text=context.bot_data["movement"].move_card(int(update.message.text)),
                 context=context)
        else:
            context.user_data["move_card"] = True
            send(chat_id=update.effective_chat.id,
                 text="Enter pair number",
                 context=context)

    @staticmethod
    def table_card(update: Update, context: CallbackContext):
        if not context.bot_data.get("movement"):
            send(chat_id=update.effective_chat.id,
                 text="Movement not set",
                 context=context)
        elif context.user_data.get("table_card"):
            context.user_data["table_card"] = False
            send(chat_id=update.effective_chat.id,
                 text=context.bot_data["movement"].table_card(int(update.message.text)),
                 context=context)
        else:
            context.user_data["table_card"] = True
            send(chat_id=update.effective_chat.id,
                 text="Enter table number",
                 context=context)

    @staticmethod
    def move_cards(update: Update, context: CallbackContext):
        if not context.bot_data.get("movement"):
            send(chat_id=update.effective_chat.id,
                 text="Movement not set",
                 context=context)
        else:
            context.bot.send_document(update.effective_chat.id, open(context.bot_data["movement"].pdf(), 'rb'))

    @staticmethod
    def freeform(update: Update, context: CallbackContext):
        text = update.message.text
        if BOARDS_RE.match(text):
            return CommandHandlers.number(update, context)
        if context.user_data.get("match_result"):
            return MatchHandlers.add_match(update, context)
        if context.user_data.get("tournament_title"):
            return CommandHandlers.title(update, context)
        if context.user_data.get("add_player"):
            return CommandHandlers.add_player(update, context)
        if context.user_data.get("update_player"):
            return CommandHandlers.update_player(update, context)
        if context.user_data.get("add_td"):
            return CommandHandlers.add_td(update, context)
        if context.user_data.get("penalty"):
            return CommandHandlers.penalty(update, context)
        if context.user_data.get('config_update'):
            return CommandHandlers.config_update(update, context)
        if context.user_data.get("names") is not None and \
                re.match('[\w ]+[-–—][\w ]+', text) or re.match('[\w ]+ [\w ]+', text):
            return CommandHandlers.names_text(update, context)
        if update.effective_chat.id > 0:
            send(update.effective_chat.id, "Unknown command", [], context)

    @staticmethod
    def names_text(update: Update, context: CallbackContext):
        if not context.bot_data.get("maxpair") or update.effective_chat.id < 0:
            return
        first = 100 * current_session(context)
        if not context.user_data["names"]:
            conn = TourneyDB.connect()
            cursor = conn.cursor()
            cursor.execute(f"Select number from names where {first} < number and number < {first + 100}")
            added = list(set([c[0] for c in cursor.fetchall()]))
            conn.close()
            all_pairs = get_valid_pairs(context)
            buttons = [l for l in all_pairs if l not in added]
            text = "Incorrect pair number.\nPlease select pair number from the list below"
            send(chat_id=update.effective_chat.id,
                 text=text,
                 reply_buttons=buttons,
                 context=context)
            return
        global ALL_PLAYERS
        conn = TourneyDB.connect()
        cursor = conn.cursor()
        pair_number = context.user_data['names']
        context.user_data["names"] = None
        found_pair_data = Players.lookup(update.message.text, ALL_PLAYERS)
        found_pair = " & ".join(p[0] for p in found_pair_data)
        rank = sum(p[1] for p in found_pair_data) / len(found_pair_data)
        rank_ru = sum(p[2] for p in found_pair_data) / len(found_pair_data)
        statement = f"""INSERT INTO names (number, partnership, rank, rank_ru)
                        VALUES({first + pair_number}, '{found_pair}', {rank}, {rank_ru}) 
                        ON CONFLICT (number) DO UPDATE 
      SET partnership = excluded.partnership,rank=excluded.rank,rank_ru=excluded.rank_ru"""
        cursor.execute(statement)
        conn.commit()
        conn.close()
        missing_players = [p[0] for p in found_pair_data if len(p) == 4]
        if CONFIG.get('store_names', True):
            text = f"Identified as {found_pair}. Players {', '.join(missing_players)} are missing. Please add them to the club by sending /addplayer command, then re-add the names" if missing_players \
                else f"Identified as {found_pair}. What's next?"
        else:
            text = f"Added {found_pair} to the tournament"
        send(chat_id=update.effective_chat.id,
             text=text,
             reply_buttons=("/names", ("/startround" if 'Swiss' in CONFIG['scoring'] else "/board")),
             context=context)

    @staticmethod
    def number(update: Update, context: CallbackContext):
        pure_number = update.message.text.split('(')[0]
        if context.user_data.get("match_result"):
            if context.user_data["match_result"]['boards'] is None:
                context.user_data["match_result"]['boards'] = int(update.message.text)
                context.user_data["match_result"]['players']['A'] = []
                send(update.effective_chat.id, 'Submit members of team A one by one', context=context)
                return
            else:
                return MatchHandlers.calculate(update, context)
            # boards will be handled below
        if len(pure_number) > 3:
            # Telegram ID
            return CommandHandlers.freeform(update, context)
        if context.user_data.get("tourney_coeff") and pure_number == '1':
            return CommandHandlers.tourney_coeff(update, context)
        first = 100 * current_session(context)
        if context.user_data.get('remove_pair'):
            return CommandHandlers.remove_pair(update, context)
        if context.user_data.get('select_session'):
            return CommandHandlers.switch_session(update, context)
        if context.user_data.get('table_card'):
            return CommandHandlers.table_card(update, context)
        if context.user_data.get('move_card'):
            return CommandHandlers.move_card(update, context)
        if context.user_data.get("penalty"):
            if context.user_data.get("penalized_pair"):
                return CommandHandlers.penalty(update, context)
            pair_number = int(update.message.text)
            context.user_data["penalized_pair"] = pair_number
            instruction = f"""
    1 for default value ({CONFIG['mp']['base_penalty']} * board MAX), or
    any other number for the respective multiple of the base penalty, or 
    40% for the specific value in %""" if CONFIG["scoring"] == "MPs" else f"""
    1 for default value ({CONFIG['imp']['base_penalty']} IMPs), or
    any other number for the respective multiple of the base penalty
"""
            send(chat_id=update.effective_chat.id,
                 text=f"Enter penalty value for pair {pair_number}. Type:{instruction}",
                 reply_buttons=[],
                 context=context)
            return
        if context.user_data.get("names", -1) == 0:
            pair_number = int(pure_number)
            context.user_data["names"] = pair_number
            send(chat_id=update.effective_chat.id,
                 text=f"Enter names for pair #{pair_number}",
                 reply_buttons=[],
                 context=context)
            return
        if context.user_data.get("rounds"):
            context.user_data["rounds"] = False
            CONFIG["rounds"] = int(update.message.text)
            if 'wiss' in CONFIG['scoring']:
                send(chat_id=update.effective_chat.id,
                     text=f"""The number of rounds is set to {update.message.text}.
    Enter /names, then use /startround or /restartswiss to proceed.""",
                     reply_buttons=[], context=context)
            else:
                try:
                    CommandHandlers.init_movement(update, context)
                except MovementError as e:
                    send(chat_id=update.effective_chat.id,
                         text=str(e),
                         reply_buttons=['/names', '/board'],
                         context=context)
            return
        if context.bot_data["maxboard"] and context.bot_data["maxpair"]:
            conn = TourneyDB.connect()
            cursor = conn.cursor()
            cursor.execute(f"Select * from boards where number={first + int(pure_number)}")
            brd = cursor.fetchall()
            if brd or CONFIG.get('no_hands', False):
                brd = brd[0] if brd else None
                if context.user_data.get("view_board"):
                    if not brd:
                        send(update.effective_chat.id, "Board not found", [], context)
                    else:
                        number = int(pure_number) + first
                        cursor.execute(
                            f"select MOD(ns, 100), MOD(ew, 100), contract, declarer, lead, result, score from protocols where number={number}")
                        board_results_raw = cursor.fetchall()
                        if not is_director(update) and context.bot_data["maxpair"] // 2 > len(board_results_raw):
                            send(chat_id=update.effective_chat.id,
                                 text="Board is still in play and you don't have director permissions",
                                 context=context)
                            context.user_data["view_board"] = False
                            return
                        view_hands(brd, board_results_raw, update, context)
                    context.user_data["view_board"] = False
                    return
                elif context.user_data.get("remove_board"):
                    context.user_data["remove_board"] = False
                    board_number = int(pure_number.strip())
                    cursor.execute(f"delete from boards where number={board_number + first}")
                    conn.commit()
                    conn.close()
                    send(chat_id=update.effective_chat.id,
                         text=f"board #{board_number} has been removed",
                         reply_buttons=[], context=context)
                    return
                # The line below is required!
                context.user_data["board"] = Board(number=int(pure_number) + first)
                conn.close()
                return result(update, context)
            context.user_data["board"] = Board()
            context.user_data["board"].number = first + int(pure_number)
            send(chat_id=update.effective_chat.id,
                 text="Enter N hand",
                 reply_buttons=[],
                 context=context)

            send(chat_id=update.effective_chat.id,
                        text="♠ \n♥ \n♦ \n♣ ",
                        reply_buttons=context.user_data["board"].get_remaining_cards(),
                        context=context)
            conn.close()
        elif context.bot_data["maxboard"]:
            context.bot_data["maxpair"] = int(update.message.text)
            if "Swiss" in CONFIG.get("scoring"):
                context.user_data["rounds"] = True
                send(chat_id=update.effective_chat.id,
                     text="Enter number of rounds",
                     reply_buttons=list(range(1, context.bot_data["maxpair"])),
                     context=context)
            else:
                try:
                    CommandHandlers.init_movement(update, context)
                except MovementError as e:
                    if 'Set' in str(e):
                        context.user_data["rounds"] = True
                        rounds = str(e).split(': ')[-1].split(',')
                        send(chat_id=update.effective_chat.id,
                             text=str(e),
                             reply_buttons=rounds,
                             context=context)
                    else:
                        send(chat_id=update.effective_chat.id,
                             text=str(e),
                             reply_buttons=['/names', '/board'],
                             context=context)
                except Exception as e:
                    send(chat_id=update.effective_chat.id,
                         text=str(e),
                         reply_buttons=[],
                         context=context)
        else:
            context.bot_data["maxboard"] = int(update.message.text)
            session_index = context.bot_data.get('session_index', 0)
            if session_index:
                conn = TourneyDB.connect()
                cursor = conn.cursor()
                fields = ','.join(f'{h}{s}' for h in hands for s in SUITS)
                first = context.bot_data['first_board']
                cursor.execute(f'select "number",{fields} from boards where and {first} <= "number" and "number" < 100 order by "number"')

                leftover_boards = cursor.fetchall()
                if len(leftover_boards) >= context.bot_data["maxboard"]:
                    modulo = CONFIG.get("sessions", {}).get("modulo")
                    for i, n in enumerate(leftover_boards):
                        new_number = (n[0] - first) % modulo + 1 if modulo else i + 1
                        cursor.execute(f"""insert into boards ("number", {fields}) 
                                                        VALUES({new_number + 100 * session_index}, '{"','".join(n[1:])}')""")
            send(chat_id=update.effective_chat.id,
                 text="Enter the number of pairs",
                 reply_buttons=[],
                 context=context)

    @staticmethod
    def init_movement(update: Update, context: CallbackContext):
        boards = context.bot_data["maxboard"]
        context.bot_data["movement"] = Movement(boards, context.bot_data["maxpair"],
                                                current_session(context),
                                                rounds=CONFIG.get('rounds', 0))
        tables = context.bot_data["movement"].tables
        rounds = context.bot_data["movement"].rounds
        CONFIG['rounds'] = rounds
        send(chat_id=update.effective_chat.id,
             text=f"Found movement for {boards} boards, {tables} tables, {rounds} rounds",
             reply_buttons=['/names', '/board'],
             context=context)

    @staticmethod
    def ok(update: Update, context: CallbackContext):
        board = context.user_data["board"]
        hand = context.user_data["currentHand"].replace("10", "T").replace(' ', '')
        if len(hand) != 20:
            send(chat_id=update.effective_chat.id,
                 text=f"Hand has incorrect number of cards ({len(hand) - 7}). Try again",
                 reply_buttons=(),
                 context=context)
            send(chat_id=update.effective_chat.id,
                text=context.user_data["currentHand"],
                reply_buttons=board.get_remaining_cards(),
                context=context)
            return
        board.set_hand(context.user_data["currentHand"])
        try:
            context.user_data["board"].is_valid()
        except RepeatingCardsException as e:
            send(chat_id=update.effective_chat.id,
                 text=str(e),
                 context=context)
            return CommandHandlers.restart(update, context)
        if board.current_hand == "w":
            w_hand = board.get_w_hand()
            w = w_hand.replace("T", "10")
            send(chat_id=update.effective_chat.id,
                 text=f"W hand should be: {w}",
                 reply_buttons=[],
                 context=context)
            if board.number:
                board.save()
                if user_allowed_boards := context.bot_data.get("allowed_boards").get(update.effective_chat.id):
                    user_allowed_boards.append(board.number)
                else:
                    context.bot_data.get("allowed_boards")[update.effective_chat.id] = [board.number]
                send(chat_id=update.effective_chat.id,
                     text=f"Board {board.number % 100} is saved",
                     reply_buttons=("/board", "/result", "/restart"),
                     context=context)
            else:
                send(chat_id=update.effective_chat.id,
                     text=f"Select board number first",
                     context=context)

        else:
            seat = board.current_hand.upper()
            send(chat_id=update.effective_chat.id,
                 text=f"Enter {seat} hand",
                 reply_buttons=[],
                 context=context)

            send(chat_id=update.effective_chat.id,
                                                    text="♠ \n♥ \n♦ \n♣ ",
                                                    reply_buttons=board.get_remaining_cards(),
                                                    context=context)

    @staticmethod
    def cancel(update: Update, context: CallbackContext):
        board = context.user_data["board"]
        board.unset_hand()
        hand = board.current_hand.upper()
        send(chat_id=update.effective_chat.id,
             text=f"Enter {hand} hand again",
             reply_buttons=[],
             context=context)
        send(chat_id=update.effective_chat.id,
                                                text="♠ \n♥ \n♦ \n♣",
                                                reply_buttons=board.get_remaining_cards(),
                                                context=context)

    @staticmethod
    def restart(update: Update, context: CallbackContext):
        board = context.user_data["board"]
        number = board.number
        context.user_data["board"] = Board()
        context.user_data["board"].number = number
        board = context.user_data["board"]

        send(chat_id=update.effective_chat.id,
             text="Enter N hand again",
             reply_buttons=[],
             context=context)
        send(chat_id=update.effective_chat.id,
                                                text="♠ \n♥ \n♦ \n♣ ",
                                                reply_buttons=board.get_remaining_cards(),
                                                context=context)

    @staticmethod
    def restart_swiss(update: Update, context: CallbackContext):
        """Clears the 'already played' matrix before the next round"""
        context.bot_data["movement"].restart()
        return CommandHandlers.start_round(update, context)

    @staticmethod
    def correct_swiss(update: Update, context: CallbackContext):
        """Only used in Swiss mode. Lists all existing boards"""
        last_played_board = context.bot_data['movement'].round * context.bot_data["maxboard"] // CONFIG["rounds"]
        send(chat_id=update.effective_chat.id,
             text="Enter board number",
             reply_buttons=list(range(1, last_played_board + 1)),
             context=context)

    @staticmethod
    def start_round(update: Update, context: CallbackContext):
        """
        Only used in Swiss mode
        """
        if 'Swiss' not in CONFIG['scoring']:
            send(chat_id=update.effective_chat.id,
                 text="Unsuitable command, use it with swiss scoring",  reply_buttons=[], context=context)
        else:
            if not context.bot_data.get("movement"):
                context.bot_data["movement"] = SwissMovement(context.bot_data["maxpair"])
            chat_id = update.message.chat_id
            try:
                context.bot.deleteMessage(chat_id, update.message.message_id)
            except Exception:
                pass
            pairings = context.bot_data["movement"].start_round()
            send(chat_id=update.effective_chat.id,
                 text=pairings,
                 reply_buttons=[], context=context)

    @staticmethod
    def end_round(update: Update, context: CallbackContext):
        """
        Only used in Swiss mode
        """
        if 'Swiss' not in CONFIG['scoring']:
            send(chat_id=update.effective_chat.id,
                 text="Unsuitable command, use it with swiss scoring",  reply_buttons=[], context=context)
        else:
            chat_id = update.message.chat_id
            try:
                context.bot.deleteMessage(chat_id, update.message.message_id)
            except Exception:
                pass
            header = send(chat_id, "Calculating results...", None, context)
            boards = context.bot_data["maxboard"]
            context.bot_data['result_getter'] = ResultGetter(boards=boards,
                                                             pairs=context.bot_data["maxpair"],
                                                             movement=context.bot_data.get('movement'))
            paths = context.bot_data['result_getter'].process()
            context.bot_data['movement'].totals = [
                t[1] for t in sorted(context.bot_data['result_getter'].totals, key=lambda x: x[0])
            ]
            if context.bot_data['maxpair'] % 2:
                context.bot_data['movement'].totals.append(0)
            for path in paths:
                context.bot.send_document(chat_id, open(path, 'rb'))
                os.remove(path)
            context.bot.editMessageText(f'Standings after round #{context.bot_data["movement"].round}:', chat_id,
                                        header.message_id)

    @staticmethod
    def testend(update: Update, context: CallbackContext):
        chat_id = update.message.chat_id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough rights to see tourney results", context=context)
            return
        try:
            context.bot.deleteMessage(chat_id, update.message.message_id)
        except Exception:
            pass
        header = send(chat_id, "Calculating results...", None, context)
        if 'BOT_TOKEN' in os.environ:
            city_en = CITIES_LATIN.get(CONFIG["city"], transliterate.translit(CONFIG["city"], 'ru'))
            path = TourneyDB.dump(city_en) if 'CURRENT_TOURNEY' in os.environ else db_path
            context.bot.send_document(update.message.from_user.id, open(path, 'rb'))
            if 'CURRENT_TOURNEY' in os.environ:
                os.remove(path)
        try:
            context.bot_data['result_getter'] = ResultGetter(boards=context.bot_data["maxboard"],
                                                             pairs=context.bot_data["maxpair"],
                                                             movement=context.bot_data.get('movement'))
            context.bot_data['result_getter'].debug = True
            if context.bot_data.get('current_session'):
                context.bot_data['result_getter'].current_session = context.bot_data.get('current_session')
            paths = context.bot_data['result_getter'].process()
            if context.bot_data['result_getter'].suspicious_result_list:
                send(update.message.from_user.id, context.bot_data['result_getter'].suspicious_results_text(),
                     context=context)
            context.bot.editMessageText(f"Tournament results:", chat_id, header.message_id)
            for path in paths:
                context.bot.send_document(chat_id, open(path, 'rb'))
                os.remove(path)
        except Exception as e:
            send(chat_id=chat_id, text=f"Result getter failed with error: {e}", context=context)
            raise

    @staticmethod
    def add_raw_result(update: Update, context: CallbackContext):
        chat_id = update.message.chat_id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough permissions to see tourney results", context=context)
            return
        if not AM:
            send(chat_id=chat_id, text="This command is not applicable to your club", context=context)
            return
        if not CONFIG.get('site_db_autoadd'):
            send(chat_id=chat_id, text="Enable site_db_autoadd in config first", context=context)
        context.bot_data['result_getter'] = ResultGetter(boards=context.bot_data["maxboard"],
                                                         pairs=context.bot_data["maxpair"])
        tournament_id = context.bot_data['result_getter'].process_ranks()
        url = CONFIG.get('tournament_url').format(tournament_id)
        context.bot.send_message(text=f"Tournament results: {url}", chat_id=chat_id)

    @staticmethod
    def end(update: Update, context: CallbackContext):
        chat_id = update.message.chat_id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough permissions to see tourney results", context=context)
            return
        try:
            context.bot.deleteMessage(chat_id, update.message.message_id)
        except Exception:
            pass
        header = send(chat_id, "Calculating results...", None, context)
        if 'BOT_TOKEN' in os.environ:
            city_en = CITIES_LATIN.get(CONFIG["city"], transliterate.translit(CONFIG["city"], 'ru'))
            path = TourneyDB.dump(city_en, movement=context.bot_data.get('movement')) if 'CURRENT_TOURNEY' in os.environ else db_path
            context.bot.send_document(update.message.from_user.id, open(path, 'rb'))
            if 'CURRENT_TOURNEY' in os.environ:
                os.remove(path)
        try:
            context.bot_data['result_getter'] = ResultGetter(boards=context.bot_data["maxboard"],
                                                             pairs=context.bot_data["maxpair"],
                                                             movement=context.bot_data.get('movement'))
            if context.bot_data.get('current_session') is not None:
                context.bot_data['result_getter'].current_session = context.bot_data.get('current_session')
                paths = context.bot_data['result_getter'].process_multisession()
            else:
                paths = context.bot_data['result_getter'].process()
            if CONFIG.get('site_db_autoadd') and context.bot_data.get('current_session') is None:
                conn = Players.connect()
                cursor = conn.cursor()
                cursor.execute(
                    "select tournament_id, boards, players, date from tournaments order by tournament_id desc"
                )
                last = cursor.fetchone()
                last_date = datetime.datetime(year=last[3].year, month=last[3].month, day=last[3].day)

                if last[1] == context.bot_data["maxboard"] and last[2] == context.bot_data["maxpair"]\
                        and last_date < datetime.datetime.now() - datetime.timedelta(hours=7):
                    cursor.execute(f"select ns,es,ws,ss from boards where tournament_id={last[0]} order by number")
                    old_boards = [".".join(c) for c in cursor.fetchall()]
                    tourney_conn = TourneyDB.connect()
                    tourney_cursor = tourney_conn.cursor()
                    tourney_cursor.execute("select ns,es,ws,ss from boards order by number")
                    new_boards = [".".join(map(lambda s: s.upper().replace('T', '10'), b))
                                  for b in tourney_cursor.fetchall()]
                    correction = len([b for b in old_boards if b in new_boards]) > last[1] * 5 // 6
                    tourney_conn.close()
                else:
                    correction = False
                conn.close()
                tournament_id = context.bot_data['result_getter'].save(correction=correction)
                if context.bot_data['result_getter'].suspicious_result_list:
                    send(update.message.from_user.id, context.bot_data['result_getter'].suspicious_results_text(),
                         context=context)
                url = CONFIG.get('tournament_url').format(tournament_id)
                context.bot.editMessageText(f"Tournament results: {url}", chat_id, header.message_id)
            elif CONFIG.get('site_ftp_upload'):
                url = CONFIG.get('tournament_url')
                parsed_url = urlparse(url)
                from ftplib import FTP
                with FTP(parsed_url.hostname) as ftp, open(paths[0], 'rb') as file:
                    ftp.cwd(parsed_url.path)
                    ftp.storbinary(f'STOR tour.shtml', file)
                context.bot.editMessageText(f"Tournament results: {url}", chat_id, header.message_id)
            else:
                context.bot.editMessageText(f"Tournament results:", chat_id, header.message_id)
                for path in paths:
                    context.bot.send_document(chat_id, open(path, 'rb'))
                    os.remove(path)
        except IncompleteTournamentData as e:
            send(chat_id=chat_id, text=f"Missing tournament data:\n{e}", context=context)
        except Exception as e:
            send(chat_id=chat_id, text=f"Result getter failed with error: {e}", context=context)
            raise

        if 'BOT_TOKEN' in os.environ and 'CURRENT_TOURNEY' not in os.environ:
            current_dir = os.getcwd()
            new_dir = f"{current_dir}/{date}"
            shutil.rmtree(new_dir)

    @staticmethod
    def excel(update: Update, context: CallbackContext):
        city_en = CITIES_LATIN.get(CONFIG["city"], transliterate.translit(CONFIG["city"], 'ru'))
        out_path = f'{city_en}.xlsx'
        indices = {
            'Ереван': 475, "Воронеж": 300, 'Ессентуки': 450, "Ижевск": 550
        }
        wb = Workbook()
        wb.create_sheet("Names")
        wb.remove(wb['Sheet'])
        try:
            conn = TourneyDB.connect()
            cursor = conn.cursor()
            cursor.execute(f"select number, partnership, rank_ru from names")
            players = cursor.fetchall()
            player_data = Players.get_players('id_ru,full_name')
            for col, value in enumerate(('Pair number', 'Name1', 'ID1', 'Name2', 'ID2', 'Rank')):
                wb['Names'].cell(1, col + 1).value = value
            for row, p in enumerate(players):
                wb['Names'].cell(row + 2, 1).value = p[0] + indices.get(CONFIG['city'], 0)
                for name_index, player_name in enumerate(p[1].split(' & ')[:2]):
                    wb['Names'].cell(row + 2, 2 * name_index + 2).value = revert_name(player_name)
                    found = [player[0] for player in player_data if player[1].split() == player_name.split()]
                    if found:
                        wb['Names'].cell(row + 2, 2 * name_index + 3).value = found[0]
                    else:
                        wb['Names'].cell(row + 2, 2 * name_index + 3).fill = PatternFill(start_color="FFC7CE",
                                                                                         end_color="FFC7CE",
                                                                                         fill_type="solid")
                wb['Names'].cell(row + 2, 6).value = p[2]

            wb.create_sheet("Protocols")
            for col, value in enumerate(('Board', 'NS', 'EW', 'Contract', 'Lead', 'Result', 'Score')):
                wb['Protocols'].cell(1, col + 1).value = value
            columns = 'number,ns,ew,contract,declarer,lead,result,score'
            cursor.execute(f"select {columns} from protocols where number > 0 order by number, score")
            results = cursor.fetchall()
            for row, r in enumerate(results):
                wb['Protocols'].cell(row + 2, 1).value = r[0]
                wb['Protocols'].cell(row + 2, 2).value = r[1] + indices.get(CONFIG['city'], 0)
                wb['Protocols'].cell(row + 2, 3).value = r[2] + indices.get(CONFIG['city'], 0)
                wb['Protocols'].cell(row + 2, 4).value = (r[3] + r[4]).upper()
                wb['Protocols'].cell(row + 2, 5).value = r[5]
                wb['Protocols'].cell(row + 2, 6).value = r[6]
                wb['Protocols'].cell(row + 2, 7).value = r[7]
            wb.save(out_path)
            context.bot.send_document(update.message.from_user.id, open(out_path, 'rb'))
        except:
            raise
        finally:
            conn.close()
            if os.path.exists(out_path):
                os.remove(out_path)

    @staticmethod
    def remove_board(update: Update, context: CallbackContext):
        chat_id = update.message.chat_id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough rights to remove boards", context=context)
            return

        context.user_data["remove_board"] = True
        send(chat_id=update.effective_chat.id,
             text=f"Enter the number of board to remove",
             reply_buttons=[], context=context)

    @staticmethod
    def remove_pair(update: Update, context: CallbackContext):
        chat_id = update.message.chat_id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough permissions to remove pairs", context=context)
            return
        if context.user_data.get("remove_pair", False):
            conn = TourneyDB.connect()
            cursor = conn.cursor()
            first = 100 * current_session(context)
            pair_to_remove = int(update.message.text)
            move_up = CONFIG.get('no_first_pair', False)

            cursor.execute(f"delete from names where number={pair_to_remove + first}")
            if move_up:
                for i in range(pair_to_remove - 1, 0, -1):
                    cursor.execute(f"""update names set number = {first + i + 1} where number={first + i}""")
            else:
                for i in range(pair_to_remove + 1, context.bot_data["maxpair"] + 1):
                    cursor.execute(f"""update names set number = {first + i - 1} where number={first + i}""")
            context.user_data["remove_pair"] = False
            cursor.execute(f"select number,partnership from names where {first} < number and number < {first + 100} "
                           f"order by number")
            found_pairs = [f'{p[0]}: {p[1]}' for p in cursor.fetchall()]
            send(chat_id=update.effective_chat.id,
                 text=f"""Pair #{pair_to_remove} has been removed, list of players:""",
                 context=context)
            send(chat_id=update.effective_chat.id,
                 text='\n'.join(found_pairs),
                 context=context)
            context.bot_data["maxpair"] -= 1
            conn.commit()
            conn.close()
            if context.bot_data["maxpair"] % 2:
                try:
                    CommandHandlers.init_movement(update, context)
                except MovementError as e:
                    send(chat_id=update.effective_chat.id,
                         text=str(e),
                         reply_buttons=['/names', '/board'],
                         context=context)
            else:
                send(chat_id=update.effective_chat.id,
                     text=f"Create new session to change the number of boards",
                     reply_buttons=['/session'],
                     context=context)
            return
        else:
            context.user_data["remove_pair"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter the number of pair to remove",
                 reply_buttons=list(range(1, context.bot_data['maxpair'] + 1)), context=context)


    @staticmethod
    def view_board(update: Update, context: CallbackContext):
        chat_id = update.message.chat_id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough permissions to see tourney boards", context=context)
            return
        context.user_data["view_board"] = True
        send(chat_id=chat_id, text=f"Enter board number",
             reply_buttons=range(1, context.bot_data["maxboard"] + 1), context=context)

    @staticmethod
    def tourney_coeff(update: Update, context: CallbackContext):
        if context.user_data.get("tourney_coeff"):
            context.user_data["tourney_coeff"] = False
            new_coeff = float(update.message.text)
            CONFIG["tourney_coeff"] = new_coeff
            if new_coeff == 1:
                # Updating ranks for national championship
                try:
                    conn = TourneyDB.connect()
                    cursor = conn.cursor()
                    cursor.execute(f"select number, partnership, rank_ru from names")
                    names = set()
                    for partnership in cursor.fetchall():
                        found_pair = Players.lookup(partnership[1], ALL_PLAYERS)
                        new_rank = sum(p[2] for p in found_pair) / len(found_pair)
                        if abs(new_rank - partnership[2]) > 0.05:
                            cursor.execute(f'update names set rank_ru={new_rank} where number={partnership[0]}')
                            names.add(partnership[0] % 100)
                    names = sorted(list(names))
                    if names:
                        conn.commit()
                except:
                    names = []
                finally:
                    conn.close()
                send(chat_id=update.effective_chat.id,
                     text=f"New tournament coefficient is {new_coeff}." +
                          (f" Ranks are updated for pairs {', '.join(map(str, names))}" if names else ''),
                     reply_buttons=[], context=context)
            else:
                send(chat_id=update.effective_chat.id,
                     text=f"New tournament coefficient is {new_coeff}",
                     reply_buttons=[], context=context)
        else:
            context.user_data["tourney_coeff"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter new tourney coeff (0.25 is the default)",
                 reply_buttons=[], context=context)

    @staticmethod
    def penalty(update: Update, context: CallbackContext):
        if context.user_data.get("penalized_pair", 0) > 0:
            context.user_data["penalty"] = False
            pair_number = context.user_data.get("penalized_pair")
            context.user_data["penalized_pair"] = False
            scoring = CONFIG["scoring"]
            pairs = context.bot_data["maxpair"]
            raw_penalty = update.message.text
            max_per_board = pairs - 2 - pairs % 2
            if raw_penalty.endswith('%'):
                mp = max_per_board * float(raw_penalty.strip('%')) / 100
            else:
                # number of base penalties
                try:
                    if scoring == "MPs":
                        mp = float(raw_penalty) * max_per_board * CONFIG['mp']['base_penalty']
                    else:
                        mp = float(raw_penalty) * CONFIG['imp']['base_penalty']
                except Exception:
                    send(chat_id=update.effective_chat.id, text="Cannot parse penalty value, try /penalty again",
                         reply_buttons=[], context=context)
                    return
            conn = TourneyDB.connect()
            cursor = conn.cursor()
            first = 100 * current_session(context)
            cursor.execute(
                f"select penalty from names where penalty is not NULL and penalty > 0 and number={pair_number + first}")
            result = cursor.fetchone()
            old_penalty = result[0] or 0 if result else 0
            cursor.execute(f"update names set penalty={old_penalty + mp} where number={pair_number + first}")
            rounded_mp = round(mp, 2)
            send(chat_id=update.effective_chat.id,
                 text=f"""Pair #{pair_number} has been penalized by {rounded_mp} {scoring}
Total penalty: {old_penalty + mp} {scoring}""",
                 reply_buttons=[], context=context)
            conn.commit()
            conn.close()
        else:
            context.user_data["penalty"] = True
            context.user_data["penalized_pair"] = 0
            all_pairs = get_valid_pairs(context)
            send(chat_id=update.effective_chat.id,
                 text=f"Enter pair number to apply penalty to",
                 reply_buttons=all_pairs, context=context)

    @staticmethod
    def add_player(update: Update, context: CallbackContext):
        if context.user_data.get("add_player"):
            context.user_data["add_player"] = False
            if AM:
                first, last, gender, rank, rank_ru = update.message.text.rsplit(" ", 4)
            else:
                first, last, gender, rank_ru = update.message.text.rsplit(" ", 3)
                rank = 0
            message = Players.add_new_player(first, last, gender, rank or 0, rank_ru)
            global ALL_PLAYERS
            ALL_PLAYERS = Players.get_players()
            send(update.effective_chat.id, f"Added player {first} {last}. {message}", [], context)
        else:
            context.user_data["add_player"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter space-separated first name, last name, gender{', rank' * AM}, rank RU",
                 reply_buttons=[], context=context)

    @staticmethod
    def update_player(update: Update, context: CallbackContext):
        if context.user_data.get("update_player"):
            context.user_data["update_player"] = False
            text = update.message.text
            if AM:
                if text.count(' ') == 2:
                    last, rank, rank_ru = text.split(' ')
                    first = None
                else:
                    first, last, rank, rank_ru = text.split(' ')
            else:
                rank = None
                if text.count(' ') == 1:
                    first = None
                    last, rank_ru = text.split(' ')
                else:
                    first, last, rank_ru = text.split(' ')
            Players.update(last, first_name=first, rank=rank, rank_ru=rank_ru)
            global ALL_PLAYERS
            ALL_PLAYERS = Players.get_players()
            send(update.effective_chat.id, "Updated player", [], context)
        else:
            context.user_data["update_player"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter space-separated values: last name/full name{', rank' * AM}, rank RU",
                 reply_buttons=[], context=context)

    @staticmethod
    def remove_player(update: Update, context: CallbackContext):
        if context.user_data.get("remove_player"):
            context.user_data["remove_player"] = False
            text = update.message.text
            first, last = text.rsplit(' ', 2)
            Players.remove(last, first_name=first)
            global ALL_PLAYERS
            ALL_PLAYERS = Players.get_players()
            send(update.effective_chat.id, "Updated player", [], context)
        else:
            context.user_data["remove_player"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter space-separated values: last name/full name",
                 reply_buttons=[], context=context)

    @staticmethod
    def list_players(update: Update, context: CallbackContext):
        players_list = Players.get_players('full_name,rank_ru' + ',rank' * AM)
        send(chat_id=update.effective_chat.id,
             text=f"Name\trank\n" + '\n'.join("\t".join(map(str, p)) for p in players_list),
             reply_buttons=[], context=context)

    @staticmethod
    def get_boards_only(update: Update, context: CallbackContext):
        chat_id = update.message.chat_id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough peermissions to see tourney boards", context=context)
            return
        path = ResultGetter(boards=context.bot_data["maxboard"], pairs=context.bot_data["maxpair"]).boards_only()
        context.bot.send_document(chat_id, open(path, 'rb'))

    @staticmethod
    def td_list(update: Update, context: CallbackContext):
        send(chat_id=update.effective_chat.id, text=", ".join(DIRECTORS), context=context)

    @staticmethod
    def add_td(update: Update, context: CallbackContext):
        chat_id = update.effective_chat.id
        if not is_director(update):
            send(chat_id=chat_id, text="You don't have enough rights to add TDs", context=context)
            return

        if context.user_data.get("add_td"):
            context.user_data["add_td"] = False
            DIRECTORS.add(update.message.text)
            send(chat_id=chat_id, text="The following players have TD rights:" + ", ".join(DIRECTORS), context=context)
        else:
            context.user_data["add_td"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter nickname or Telegram ID of the new TD",
                 reply_buttons=[], context=context)

    @staticmethod
    def store(update: Update, context: CallbackContext):
        context.bot_data['result_getter'].save()

    @staticmethod
    def correct(update: Update, context: CallbackContext):
        context.bot_data['result_getter'].save(correction=True)

    @staticmethod
    def load_db(update: Update, context: CallbackContext):
        CONFIG["load_db"] = True
        send(update.effective_chat.id, "Forward/drag .db file and send", None, context)

    @staticmethod
    def custom_movement(update: Update, context: CallbackContext):
        context.bot_data["movement"] = None
        send(chat_id=update.effective_chat.id,
             text="No movement is used",
             reply_buttons=[], context=context)

    @staticmethod
    def toggle_hands(update: Update, context: CallbackContext):
        CONFIG['no_hands'] = not CONFIG['no_hands']
        send(chat_id=update.effective_chat.id,
             text="Hands are turned " + ("off" if CONFIG['no_hands'] else "on"),
             reply_buttons=[], context=context)

    @staticmethod
    def mitchell(update: Update, context: CallbackContext):
        CONFIG['is_mitchell'] = True
        send(chat_id=update.effective_chat.id,
             text="Movement is set to Mitchell",
             reply_buttons=[], context=context)

    @staticmethod
    def howell(update: Update, context: CallbackContext):
        CONFIG['is_mitchell'] = False
        send(chat_id=update.effective_chat.id,
             text="Movement is set to Howell",
             reply_buttons=[], context=context)

    @staticmethod
    def barometer(update: Update, context: CallbackContext):
        CONFIG['is_barometer'] = not CONFIG.get('is_barometer')
        send(chat_id=update.effective_chat.id,
             text="Barometer is turned " + ("on" if CONFIG['is_barometer'] else "off"),
             reply_buttons=[], context=context)


    @staticmethod
    def rounds(update: Update, context: CallbackContext):
        if not context.user_data.get("rounds"):
            context.user_data["rounds"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter the number of rounds",
                 reply_buttons=[str(i + 1) for i in range(2, context.bot_data['maxpair'] - 1)], context=context)

    @staticmethod
    def title(update: Update, context: CallbackContext):
        if context.user_data.get("tournament_title"):
            context.user_data["tournament_title"] = False
            CONFIG["tournament_title"] = update.message.text
            send(chat_id=update.effective_chat.id,
                 text=f"Set tournament title to {update.message.text}",
                 reply_buttons=[], context=context)
        else:
            context.user_data["tournament_title"] = True
            send(chat_id=update.effective_chat.id,
                 text=f"Enter new tournament title",
                 reply_buttons=[], context=context)

    @staticmethod
    def monthly_report(update: Update, context: CallbackContext):
        send(chat_id=update.effective_chat.id,
             text=Players.monthly_report(),
             reply_buttons=[], context=context)

    @staticmethod
    def donate(update: Update, context: CallbackContext):
        context.bot.send_message(chat_id=update.effective_chat.id,
                                 text="""Beer, coffee, and cloud servers don't grow on trees. 
If you want to support the developers, click the link below (RU) or use VISA card transfer: 
RU card: https://www.tinkoff.ru/cf/3mG1Qe6fN3q
AM card: `4318 2700 0032 4697`
BTC: `bitcoin:bc1q83p77e9zqe3ju5ew4crejm0f9lf9grzx0mu6nh`""", parse_mode=ParseMode.MARKDOWN)

    @staticmethod
    def help_command(update: Update, context: CallbackContext):
        text = Template(open("templates/help.html").read()).render({
            'AM': AM,
            'is_director': is_director(update),
            'swiss': "wiss" in CONFIG["scoring"]
        })
        send(chat_id=update.message.chat_id, text=text, context=context)

    @staticmethod
    def manual(update: Update, context: CallbackContext):
        send(chat_id=update.message.chat_id,
             text="https://telegra.ph/Instrukciya-dlya-telegram-bota-dlya-provedeniya-turnira-po-bridzhu-04-24",
             context=context)
